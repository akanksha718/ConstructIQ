"""Spreadsheet / CSV parsing without pandas (openpyxl + stdlib csv)."""

from __future__ import annotations

import csv
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelParser:

    def parse(self, file_path: str) -> str:
        ext = Path(file_path).suffix.lower()

        if ext == ".csv":
            return self._parse_csv(file_path)

        return self._parse_xlsx(file_path)

    @staticmethod
    def _rows_to_markdown(rows: list[list[str]]) -> str:
        rows = [r for r in rows if any(str(c).strip() for c in r)]
        if not rows:
            return ""

        header = rows[0]
        lines = [
            "| " + " | ".join(str(c) for c in header) + " |",
            "| " + " | ".join("---" for _ in header) + " |",
        ]
        for row in rows[1:]:
            padded = list(row) + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(str(c) for c in padded) + " |")
        return "\n".join(lines)

    def _parse_csv(self, file_path: str) -> str:
        try:
            with open(file_path, newline="", encoding="utf-8-sig") as fh:
                rows = list(csv.reader(fh))
        except Exception:
            logger.exception("Failed to read CSV %s", file_path)
            return ""
        return self._rows_to_markdown(rows)

    def _parse_xlsx(self, file_path: str) -> str:
        try:
            from openpyxl import load_workbook
        except Exception:  # pragma: no cover - dependency missing
            logger.exception("openpyxl not available")
            return ""

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception:
            logger.exception("Failed to open workbook %s", file_path)
            return ""

        sections = []
        for sheet in wb.worksheets:
            rows = [
                [("" if c is None else c) for c in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            table = self._rows_to_markdown(rows)
            if table:
                sections.append(f"# Sheet: {sheet.title}\n\n{table}")

        return "\n\n".join(sections)
